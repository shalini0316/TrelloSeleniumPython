import openpyxl
from easygui import passwordbox


class HomePageData:

    test_HomePage_data = [{"Email":"shalini"}, {"Email":"Anshika"}]

    @staticmethod
    def getTestData():

        excelList=[]
        book = openpyxl.load_workbook("C:\\Users\\BLACKPEARL COMPUTERS\\PycharmProjects\\pythonProject\\PythonSelFramework\\TestData\\testsuite.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=2).value == 'Y':
                excelDict = {}
                for j in range(1, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    excelDict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                excelList.append(excelDict)
        return excelList

    @staticmethod
    def getPassword():
        passwordList = []
        passwordDict = {}
        passwordDict["password"]=passwordbox("PASSWORD:")
        passwordList.append(passwordDict)
        return passwordList

